import openpyxl


def get_data(sheetName):
    # return [
    #     ["Delhi"],
    #     ["Dubai"]
    # ]
    workbook = openpyxl.load_workbook("..\\Excel\\testdata.xlsx")
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcol = sheet.max_column
    mainlist = []
    for i in range(2, totalrows + 1):
        insidelist = []
        for j in range(1, totalcol + 1):
            data = sheet.cell(row=i, column=j).value
            insidelist.insert(j, data)
        mainlist.insert(i, insidelist)
    return mainlist