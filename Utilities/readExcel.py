import openpyxl

workbook = openpyxl.load_workbook("..//Excel//testdata.xlsx")
sheet = workbook["logintest"]
totalrows = sheet.max_row
totalcol = sheet.max_column

print("Columns : ", str(totalcol), "Rows : ", str(totalrows))

valueatcell = sheet.cell(row=2,column=2).value
print(valueatcell)